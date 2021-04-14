"""
이름 : 박민지
날짜 : 2021/04/14
내용 : 파이썬 외부 패키지 설치, 특수파일 실습 교재 p239
"""
from openpyxl import Workbook
import pandas as pd

# Excel 파일 읽기
exam = pd.read_excel('./data/exam.xlsx')
print(exam)

# Excel 파일쓰기
# 새로운 엑셀파일 생성
workbook = Workbook()

# 현재 sheet 활성화
sheet = workbook.active

# 데이터 입력
sheet['A1'] = 'A1셀'
sheet.append([1, 2, 3])
sheet.append(['김유신', '김춘추', '장보고', '강감찬', '이순신'])
sheet.cell(5, 5, '5x5 데이터')
sheet.cell(6, 2, '6x2 데이터')

# 파일 저장/닫기
workbook.save('C:/Users/501/Desktop/Sample.xlsx')
workbook.close()

print('Excel 파일생성 완료')